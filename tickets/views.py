import openpyxl
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.models import User
from django.db.models import Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.views.generic import (
    ListView,
    DetailView,
    UpdateView,
    CreateView,
)
from django.views.generic.edit import FormView
from openpyxl.utils import get_column_letter

from tickets.forms import TicketForm, TicketListFilterForm
from tickets.utils import StaffMemberRequiredMixin
from .models import Ticket
from .tasks import send_ticket_email_task


class TicketCreateView(LoginRequiredMixin, CreateView):
    model = Ticket
    template_name_suffix = "_new"
    form_class = TicketForm

    def form_valid(self, form):
        self.object = form.save(commit=False)
        self.object.author = self.request.user
        self.object.save()

        absolute_uri = self.request.build_absolute_uri("/")
        send_ticket_email_task.delay(self.object.id, absolute_uri)

        return super().form_valid(form)

    def get_success_url(self):
        return reverse("ticket_detail", kwargs={"pk": self.object.id})


class TicketEditView(StaffMemberRequiredMixin, UpdateView):
    model = Ticket
    fields = ["status", "resolved_at", "priority", "assigned_to"]
    template_name_suffix = "_edit"

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        form.fields["assigned_to"].queryset = User.objects.filter(is_staff=True)
        form.fields["status"].widget.attrs.update({"class": "form-select"})
        form.fields["priority"].widget.attrs.update({"class": "form-select"})
        form.fields["assigned_to"].widget.attrs.update({"class": "form-select"})
        form.fields["resolved_at"].widget.attrs.update(
            {"class": "form-control", "type": "datetime-local"}
        )
        return form

    def get_success_url(self):
        return reverse("ticket_detail", kwargs={"pk": self.object.id})


class TicketDetailView(LoginRequiredMixin, DetailView):
    model = Ticket
    template_name = "tickets/ticket_detail.html"
    context_object_name = "ticket"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        ticket = get_object_or_404(Ticket, id=self.kwargs["pk"])
        if (
            not self.request.user.is_staff
            and ticket.author.id is not self.request.user.id
        ):
            return render(self.request, "authorization_error.html")

        context["ticket"] = ticket
        return context


class TicketListView(LoginRequiredMixin, ListView):
    paginate_by = 10
    model = Ticket
    template_name = "tickets/ticket_list.html"
    context_object_name = "tickets"

    def get_queryset(self):
        tickets = Ticket.objects.all()

        if not self.request.user.is_staff:
            tickets = tickets.filter(author=self.request.user)

        form = TicketListFilterForm(self.request.GET)
        if form.is_valid():
            status_filter = form.cleaned_data.get("status")
            priority_filter = form.cleaned_data.get("priority")
            department_filter = form.cleaned_data.get("department")
            company_filter = form.cleaned_data.get("company")
            assigned_to_filter = form.cleaned_data.get("assigned_to")
            search_query = form.cleaned_data.get("search_query")
            initial_date = form.cleaned_data.get("initial_date")
            end_date = form.cleaned_data.get("end_date")

            if status_filter:
                tickets = tickets.filter(status=status_filter)
            if priority_filter:
                tickets = tickets.filter(priority=priority_filter)
            if department_filter:
                tickets = tickets.filter(department=department_filter)
            if company_filter:
                tickets = tickets.filter(company=company_filter)
            if assigned_to_filter:
                tickets = tickets.filter(assigned_to=assigned_to_filter)
            if search_query:
                tickets = tickets.filter(
                    Q(title__icontains=search_query)
                    | Q(description__icontains=search_query)
                    | Q(status__icontains=search_query)
                    | Q(priority__icontains=search_query)
                    | Q(author__username__icontains=search_query)
                    | Q(author__first_name__icontains=search_query)
                )
            if initial_date:
                tickets = tickets.filter(created_at__gte=initial_date)
            if end_date:
                tickets = tickets.filter(created_at__lte=end_date)

        return tickets.order_by("-updated_at")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["form"] = TicketListFilterForm(self.request.GET)
        return context

    def render_to_response(self, context, **response_kwargs):
        if "download" in self.request.GET:
            tickets = context["tickets"]

            # Prepare the Excel sheet
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Tickets"

            # Define the headers for the Excel sheet
            headers = [
                "ID",
                "Title",
                "Description",
                "Company",
                "Department",
                "Status",
                "Priority",
                "Assigned To",
                "Author",
                "Created At",
                "Updated At",
                "Resolved at",
            ]
            sheet.append(headers)

            # Populate the rows with ticket data
            for ticket in tickets:
                row = [
                    ticket.id,
                    ticket.title,
                    ticket.description,
                    ticket.company.name,
                    ticket.department.name,
                    ticket.status,
                    ticket.priority,
                    ticket.assigned_to.username if ticket.assigned_to else "",
                    ticket.author.username,
                    ticket.created_at.strftime("%Y-%m-%d %H:%M"),
                    ticket.updated_at.strftime("%Y-%m-%d %H:%M"),
                    (
                        ticket.resolved_at.strftime("%Y-%m-%d %H:%M")
                        if ticket.resolved_at
                        else "---"
                    ),
                ]
                sheet.append(row)

                # Adjust column widths based on the longest text
                for col_num in range(1, len(headers) + 1):
                    max_length = 0
                    column = get_column_letter(col_num)
                    for row in sheet.iter_rows(min_col=col_num, max_col=col_num):
                        for cell in row:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(cell.value)
                            except:
                                pass
                    adjusted_width = max_length + 2
                    sheet.column_dimensions[column].width = adjusted_width

                # Adjust row height to fit the content
                for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                    for cell in row:
                        try:
                            # Check if the row needs to be resized
                            cell_value = str(cell.value)
                            if len(cell_value) > 50:  # Adjust the number as needed
                                sheet.row_dimensions[cell.row].height = (
                                    40  # You can adjust the height value
                                )
                        except:
                            pass

            # Create an HttpResponse to serve the Excel file
            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response["Content-Disposition"] = 'attachment; filename="tickets.xlsx"'

            # Save the workbook to the response
            workbook.save(response)
            return response

        return super().render_to_response(context, **response_kwargs)